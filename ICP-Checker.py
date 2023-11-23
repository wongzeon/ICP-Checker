import re
import os
import time
import hashlib
import requests
from requests import utils
import openpyxl as xl
from openpyxl.styles import Alignment

os.environ['no_proxy'] = '*'


def query_base():
    print("版本：V2.1.9 可用测试：2023-9-25\n")
    print("项目地址：https://github.com/wongzeon/ICP-Checker\n")
    while True:
        try:
            info = (
                input("请完整输入公司全称 / 域名以查询备案信息：\n\n")
                .replace(" ", "")
                .replace("https://www.", "")
                .replace("http://www.", "")
                .replace("http://", "")
            )
            # 过滤空值和特殊字符，只允许 - —《》. () 分别用于域名和公司名
            if info == "":
                raise ValueError("InputNone")
            info = re.sub("[^\\u4e00-\\u9fa5-A-Za-z0-9,-.()《》—（）]", "", info)
            input_zh = re.compile(u'[\u4e00-\u9fa5]')
            zh_match = input_zh.search(info)
            if zh_match:
                info_result = info
            else:
                # 检测是否为可备案的域名类型（类型同步日期2022/01/06）
                # TODO: 建议移除域名类型检测, 维护不及时且部分留存域名类型较为特殊, 包括 IP 备案.
                input_url = re.compile(
                    r'([^.]+)(?:\.(?:GOV\.cn|ORG\.cn|AC\.cn|MIL\.cn|NET\.cn|EDU\.cn|COM\.cn|BJ\.cn|TJ\.cn|SH\.cn|CQ\.cn|HE\.cn|SX\.cn|NM\.cn|LN\.cn|JL\.cn|HL\.cn|JS\.cn|ZJ\.cn|AH\.cn|FJ\.cn|JX\.cn|SD\.cn|HA\.cn|HB\.cn|HN\.cn|GD\.cn|GX\.cn|HI\.cn|SC\.cn|GZ\.cn|YN\.cn|XZ\.cn|SN\.cn|GS\.cn|QH\.cn|NX\.cn|XJ\.cn|TW\.cn|HK\.cn|MO\.cn|cn|REN|WANG|CITIC|TOP|SOHU|XIN|COM|NET|CLUB|XYZ|VIP|SITE|SHOP|INK|INFO|MOBI|RED|PRO|KIM|LTD|GROUP|BIZ|AUTO|LINK|WORK|LAW|BEER|STORE|TECH|FUN|ONLINE|ART|DESIGN|WIKI|LOVE|CENTER|VIDEO|SOCIAL|TEAM|SHOW|COOL|ZONE|WORLD|TODAY|CITY|CHAT|COMPANY|LIVE|FUND|GOLD|PLUS|GURU|RUN|PUB|EMAIL|LIFE|CO|FASHION|FIT|LUXE|YOGA|BAIDU|CLOUD|HOST|SPACE|PRESS|WEBSITE|ARCHI|ASIA|BIO|BLACK|BLUE|GREEN|LOTTO|ORGANIC|PET|PINK|POKER|PROMO|SKI|VOTE|VOTO|ICU|LA))',
                    flags=re.IGNORECASE)
                info_result = input_url.search(info)
                if info_result is None:
                    if info.split(".")[0] == "":
                        raise ValueError("OnlyDomainInput")
                    raise ValueError("ValidType")
                else:
                    info_result = info_result.group()
            info_data = {'pageNum': '1', 'pageSize': '40', 'serviceType': 1, 'unitName': info_result}
            return info_data
        except ValueError as e:
            if str(e) == 'InputNone' or str(e) == 'OnlyDomainInput':
                print("\n ************** 请正确输入域名 **************\n")
            else:
                print("\n*** 该域名不支持备案，请查阅：http://xn--fiq8ituh5mn9d1qbc28lu5dusc.xn--vuq861b/ ***\n")


def get_cookies():
    cookie_headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.41 Safari/537.36 Edg/101.0.1210.32'}
    err_num = 0
    while err_num < 3:
        try:
            cookie = utils.dict_from_cookiejar(
                requests.get(
                    'https://beian.miit.gov.cn/',
                    headers=cookie_headers,
                    timeout=(3.06, 27)).cookies
            )['__jsluid_s']
            return cookie
        except:
            err_num += 1
            time.sleep(3)
    return -1


def get_token():
    timeStamp = round(time.time() * 1000)
    authSecret = 'testtest' + str(timeStamp)
    authKey = hashlib.md5(authSecret.encode(encoding='UTF-8')).hexdigest()
    auth_data = {'authKey': authKey, 'timeStamp': timeStamp}
    url = 'https://hlwicpfwc.miit.gov.cn/icpproject_query/api/auth'
    try:
        t_response = requests.post(url=url, data=auth_data, headers=base_header, timeout=(3.06, 27)).json()
        token = t_response['params']['bussiness']
    except:
        return -1
    return token


def get_beian_info(info_data, token):
    domain_list = []
    info_url = 'https://hlwicpfwc.miit.gov.cn/icpproject_query/api/icpAbbreviateInfo/queryByCondition'
    base_header.update({'token': token})
    try:
        beian_info = requests.post(url=info_url, json=info_data, headers=base_header, timeout=(3.06, 27)).json()
        if not beian_info["success"]:
            print(f'请求错误: CODE {beian_info["code"]} MSG {beian_info["msg"]}')
            return domain_list
        domain_total = beian_info['params']['total']
        page_total = beian_info['params']['lastPage']
        end_row = beian_info['params']['endRow']
        info = info_data['unitName']
        print(f"\n查询对象：{info} 共有 {domain_total} 个已备案域名\n")
        for i in range(0, page_total):
            print(f"正在查询第{i + 1}页……\n")
            for k in range(0, end_row + 1):
                info_base = beian_info['params']['list'][k]
                domain_name = info_base['domain']
                domain_type = info_base['natureName']
                domain_licence = info_base['mainLicence']
                website_licence = info_base['serviceLicence']
                domain_status = info_base['limitAccess']
                domain_approve_date = info_base['updateRecordTime']
                domain_owner = info_base['unitName']
                try:
                    domain_content_approved = info_base['contentTypeName']
                    if domain_content_approved == "":
                        domain_content_approved = "无"
                except KeyError:
                    domain_content_approved = "无"
                row_data = domain_owner, domain_name, domain_licence, website_licence, domain_type, domain_content_approved, domain_status, domain_approve_date
                domain_list.append(row_data)
            info_data = {'pageNum': i + 2, 'pageSize': '40', 'serviceType': 1, 'unitName': info}
            if beian_info['params']['isLastPage'] is True:
                break
            else:
                beian_info = requests.post(info_url, json=info_data, headers=base_header, timeout=(3.06, 27)).json()
                end_row = beian_info['params']['endRow']
                time.sleep(3)
    except Exception as e:
        print(f"意外错误: {e}")
        return domain_list
    return domain_list


def data_saver(domain_list):
    """
    打印最终结果，并保存数据至Excel表格，同时调整表格格式。
    """
    # 计算需要写入表格的总行数，如果是空列表，即代表该域名没有备案信息，也有可能是获取信息失败了
    total_row = len(domain_list)
    if total_row == 1:
        total_row = 0
    elif total_row == 0:
        return print("所查域名无备案\n")
    print(f"查询结果如下:\n\n{domain_list}\n")
    # Windows获取桌面路径，将表格保存到桌面，其他系统默认保存到/home/文件夹下
    if os.name == "nt":
        import winreg
        # 用户更改过桌面路径，则需获取User Shell Folders才能获取到准确的桌面路径，否则不会保存到实际的桌面
        subkey = r'Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders'
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, subkey, 0)
        desktop_raw = str(winreg.QueryValueEx(key, "Desktop")[0])
        if desktop_raw == "%USERPROFILE%\Desktop":
            # 此时情况为用户未更改过桌面路径，则需获取系统默认路径
            subkey = r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, subkey, 0)
            desktop_raw = str(winreg.QueryValueEx(key, "Desktop")[0])
        desktop_path = desktop_raw.replace('\\', '/') + "/"
        file_path = f"{desktop_path}备案信息.xlsx"
    else:
        file_path = '/home/备案信息.xlsx'
    # 存在对应文件，则读取表格追加写入，不存在则创建，并设置表格的标题、列宽、冻结窗格、文字布局等格式
    if os.path.exists(file_path):
        wb = xl.load_workbook(file_path)
        ws = wb['备案信息']
        max_row = ws.max_row
        start = max_row + 1
        total_row = total_row + start
        after_title = 0
    else:
        wb = xl.Workbook()
        ws = wb.active
        ws.title = "备案信息"
        title_list = ['域名主办方', '域名', '备案许可证号', '网站备案号', '域名类型', '网站前置审批项', '是否限制接入',
                      '审核通过日期']
        for i in range(0, 8):
            ws.cell(1, i + 1).value = title_list[i]
        col_width = {'A': 45, 'B': 40, 'C': 22, 'D': 24, 'E': 9, 'F': 15, 'G': 13, 'H': 21}
        for k, v in col_width.items():
            ws.column_dimensions[k].width = v
        ws.freeze_panes = 'A2'
        start = 0
        after_title = 2
    # 写入查询数据
    for j in range(start, total_row + 1):
        for k in range(0, 8):
            try:
                ws.cell(j + after_title, k + 1).value = domain_list[j - start][k]
            except:
                continue
    # 垂直居中
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            ws.cell(row + 1, col + 1).alignment = Alignment(horizontal='center', vertical='center')
    try:
        wb.save(file_path)
    except PermissionError:
        print("** 备案信息登记表格已打开，无法写入文件。如需写入，请关闭文件后重新执行！ **\n")
        return -1
    print(f"查询结果保存在：{file_path}\n")
    return 'OK'


def main():
    print("获取Cookie中，速度取决于网站响应，请等待……\n")
    cookie = get_cookies()
    while True:
        info = query_base()
        try:
            global base_header
            base_header = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.41 Safari/537.36 Edg/101.0.1210.32',
                'Origin': 'https://beian.miit.gov.cn',
                'Referer': 'https://beian.miit.gov.cn/',
                'Cookie': f'__jsluid_s={cookie}'
            }
            # -1代表对应步骤失败了，不是-1则正常执行下一步
            if cookie != -1:
                token = get_token()
                print("\n获取Token，请等待……\n")
                if token != -1:
                    print("已获取到Token，查询中，速度取决于网站响应，请等待……")
                    domain_list = get_beian_info(info, token)
                    data_saver(domain_list)
                else:
                    raise ValueError("获取Token失败，如频繁失败请关闭程序后等待几分钟再试！")
            else:
                cookie = get_cookies()
                raise ValueError("获取Cookie失败，请重试！")
        except Exception as e:
            print(f'{e}\n')


if __name__ == '__main__':
    main()
