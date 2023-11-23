# -*- coding: utf-8 -*-
import os
import re
import cv2
import random
from requests import get, post
from time import sleep, time
from hashlib import md5
from base64 import b64encode, b64decode
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment

class get_require_info(object):
    """
    获取必要的Cookie、Token，以及各类的POST请求由这里的post_tool()完成，
    """
    base_header = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36 Edg/118.0.2088.69',
                'Origin': 'https://beian.miit.gov.cn', 'Referer': 'https://beian.miit.gov.cn/'}

    def __init__(self) -> None:
        self.base_url = 'https://hlwicpfwc.miit.gov.cn/icpproject_query/api/'
        self.auth_url = self.base_url + 'auth' # 返回token
        self.image_url = self.base_url + 'image/getCheckImagePoint' # 返回Secretkey、uuid
        self.check_url = self.base_url + 'image/checkImage' # 返回Sign，传值里的token是uuid
        self.query_url = self.base_url + 'icpAbbreviateInfo/queryByCondition' # 标头需要：cookie、token、sign、uuid

    def post_tool(self, url, data, json, header):
        try:
            response = post(url, data=data, json=json, headers=header, timeout=2)
            response_code = response.json()['code']
            response_msg = response.json()['success']
            if response_code == 200 and response_msg is True:
                return response
            elif response_code == 200 and response_msg is False: # 验证错误，需要重新获取验证图片
                return 'pic_check_fail'
            elif response_code == 401: # Token 过期
                print(f"{' 刷新Token ':-^57}")
                return self.get_token()
            elif response_code == (500 or 501):
                return 'server_error'
            elif response.status_code == 403: # WAF 触发，只有这个会正确返回状态码，其他的错误状态码响应的是200
                print(f"{' 触发防火墙了，过几分钟再试吧 ':*^50}")
                exit()
        except Exception as err:
            return f'error:{err}'

    def get_cookie(self):
        cookie = get(url=self.base_header['Origin'], headers={'User-Agent':self.base_header['User-Agent']}, timeout=1)
        self.base_header.update({'Cookie': cookie.headers['Set-Cookie'].split(';')[0]})

    def get_token(self):
        print(f"\n{' 获取Token中，服务器响应时快时慢 ':-^46}")
        time_stamp = round(time() * 1000)
        auth_key = md5(f'testtest{time_stamp}'.encode('utf-8')).hexdigest()
        auth_data = {'authKey': auth_key, 'timeStamp': time_stamp}
        token_request = self.post_tool(self.auth_url, auth_data, "", self.base_header)
        token_result = token_request.json()['params']['bussiness']
        self.base_header.update({'Token': token_result, 'Accept':'application/json'})
        print(f"\n{' 已获得Token ':-^56}\n")

class file_tool(object):
    """
    文件处理，主要为批量做预留，减少重复代码
    """
    def __init__(self) -> None:
        pass
        
    def file_writer(self, file_path, file_name, file, wr_mode, wr_type):
        with open(f'{file_path}/{file_name}', wr_mode) as f:
            if wr_type == 'write':
                f.write(file)
            else:
                self.file_content = f.read()
                return self.file_content

class image_tool(get_require_info, file_tool):
    """
    图片下载、处理及点选的部分，先生成client_uid，获得SecretKey，然后进行点选、加密回传以及验证，check_image()内调用前几步
    """
    def __init__(self) -> None:
        super().__init__()

    def get_position(self):
        print(f"\n{' 获取图片 ':-^55}")
        secret_text = '0123456789abcdef'
        text_list = ''.join([secret_text[random.randint(0, 15)] for i in range(36)])
        self.client_uid = 'point-' + text_list[:8] + '-' + text_list[9:13] + '-' + text_list[14:18] + '-' + text_list[19:23] + '-' + text_list[24:]
        image_base64 = super().post_tool(self.image_url, '', {'clientUid': self.client_uid}, self.base_header).json()['params']
        self.secret_key = image_base64['secretKey']
        self.uuid = image_base64['uuid']
        super().file_writer('./', 'bigImage.png', b64decode(image_base64['bigImage']), 'wb', 'write')
        super().file_writer('./', 'smallImage.png', b64decode(image_base64['smallImage']), 'wb', 'write')

    def encrypt_position(self):
        self.position = []
        self.click_num = 0
        big_image = cv2.imread('./bigImage.png', 1)
        small_image = cv2.imread('./smallImage.png', 1)
        self.combine_image = cv2.vconcat([big_image, small_image])
        print(f"\n{' 请在弹出对话框点击对应文字，误点可点击右键撤销 ':*^36}")
        cv2.imshow('Choose the recapture text', self.combine_image)
        cv2.setWindowProperty('Choose the recapture text', cv2.WND_PROP_TOPMOST, 1)
        cv2.setMouseCallback('Choose the recapture text', self.click_position)
        cv2.waitKey(0)
        if not isinstance(self.position, bytes):
            data = str(self.position).replace("'",'"').replace(' ','').encode('utf-8')
        padded_data = pad(data, block_size=128, style='pkcs7')
        cipher = AES.new(self.secret_key.encode('utf-8'), AES.MODE_ECB)
        self.encrypted_point = b64encode(cipher.encrypt(padded_data)).decode()
    
    def click_position(self, event, x_position, y_position, flags, params):
        if event == cv2.EVENT_LBUTTONDOWN:
            self.position.append({'x':x_position, 'y':y_position})
            cv2.circle(self.combine_image, (x_position, y_position), 5, (0, 255, 153), thickness = -1)
            cv2.imshow('Choose the recapture text', self.combine_image)
            self.click_num += 1
            if self.click_num == 4:
                cv2.destroyAllWindows()
        elif event == cv2.EVENT_RBUTTONDOWN:
            del self.position[self.click_num - 1]
            self.click_num = self.click_num -1
            cv2.circle(self.combine_image, (x_position, y_position), 5, (0, 0, 255), thickness = -1)
            cv2.imshow('Choose the recapture text', self.combine_image)

    def check_image(self):
        self.get_position()
        self.encrypt_position()
        os.remove('./bigImage.png')
        os.remove('./smallImage.png')
        check_data = {'token': self.uuid, 'secretKey': self.secret_key, 'clientUid': self.client_uid, 'pointJson': self.encrypted_point}
        sign_request = super().post_tool(self.check_url, '', check_data, self.base_header)
        if sign_request == 'pic_check_fail':
            return -1
        sign_result = sign_request.json()['params']['sign']
        self.base_header.update({'Sign': sign_result, 'uuid': self.uuid})

class query_info(get_require_info, file_tool):
    """
    输入过滤，只对中文输入做过滤，英文域名由于有些域名不在可备案列表内也能备案，所以不做过滤了
    """
    domain_result_list = []
    def __init__(self) -> None:
        super().__init__()

    def regular_input(self):
        print(f"{' 输入 域名/备案号/完整公司名 以查询备案信息 ':-^40}\n")
        info_input = input('查询对象：').replace(' ', '').replace('https://www.', '').replace('http://www.', '').replace('http://', '').replace('www.', '')
        if info_input == '':
             return print(f"\n{' 请输入正确域名 ':*^52}\n")
        elif '\u4e00' <= info_input <= '\u9fff':
            input_result = re.sub('[^\\u4e00-\\u9fff-A-Za-z0-9,-.()《》—（）]', "", info_input) # 中文输入混入的特殊符号，只允许：-—《》（）().
        else:
            input_result = info_input
        self.query_text = {'pageNum': 1, 'pageSize': 40, 'serviceType': 1, 'unitName': input_result} #serviceType: 1:网站 6:APP 7:小程序 8:快应用

    def query_information(self):
        print(f"\n{' 查询中 ':-^56}")
        icp_info = super().post_tool(self.query_url, '', self.query_text, self.base_header).json()['params']
        total_domains = icp_info['total']
        total_pages = icp_info['lastPage']
        end_row = icp_info['endRow']
        if total_domains == 0:
            print(f"\n{' '+ self.query_text['unitName'] + ' 没有备案信息 ':*^40}\n")
            return -1
        print(f"\n{' '+ self.query_text['unitName'] + ' 共有 ' + str(total_domains) + ' 个备案域名 ':*^40}\n")
        for i in range(total_pages):
            print(f"{' 获取第' + str(i+1) + '页信息 ':-^53}\n")
            for k in range(0, end_row + 1):
                info_base = icp_info['list'][k]
                domain_name = info_base['domain']
                domain_type = info_base['natureName']
                domain_licence = info_base['mainLicence']
                website_licence = info_base['serviceLicence']
                domain_status = info_base['limitAccess']
                domain_approve_date = info_base['updateRecordTime']
                domain_owner = info_base['unitName']
                try:
                    domain_content_approved = info_base['contentTypeName']
                    if domain_content_approved == "":
                        domain_content_approved = "无"
                except KeyError:
                    domain_content_approved = "无"
                row_data = domain_owner, domain_name, domain_licence, website_licence, domain_type, domain_content_approved, domain_status, domain_approve_date
                self.domain_result_list.append(row_data)
            self.query_text.update({'pageNum': i + 2})
            if icp_info['isLastPage'] is True:
                break
            else:
                icp_info = super().post_tool(self.query_url, '', self.query_text, self.base_header).json()['params']
                end_row = icp_info['endRow']
                sleep(3) # 翻页需要慢一点，否则大概率被WAF拦截

class excel_tool(query_info):
    """
    保存结果到Excel文件，注意执行时需关闭“备案信息.xlsx”，否则会保存失败
    """
    def __init__(self) -> None:
        if os.name == 'nt':
            from winreg import OpenKey, QueryValueEx, HKEY_CURRENT_USER
            key = r'Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders' # 修改过桌面路径的注册表
            target_key = OpenKey(HKEY_CURRENT_USER, key, 0)
            desktop_path = str(QueryValueEx(target_key, 'Desktop')[0])
            if desktop_path == r'%USERPROFILE%\Desktop':
                desktop_path = os.path.join(os.path.expanduser('~'),"Desktop")
        else:
            desktop_path = f'/home/{os.getlogin()}'
            if not os.path.exists(desktop_path):
                os.mkdir(desktop_path)
        self.excel_file_path = desktop_path.replace('\\', '/') + '/备案信息.xlsx'
        super().__init__()

    def set_format(self):
        list_row = len(self.domain_result_list)
        if os.path.exists(self.excel_file_path):
            self.work_book = load_workbook(self.excel_file_path)
            self.work_sheet = self.work_book['备案信息']
            self.start = self.work_sheet.max_row + 1 # 从文件最后一行的后一行开始追加写入
            self.total_row = list_row + self.start
            self.after_title = 0
        else:
            self.work_book = Workbook()
            self.work_sheet = self.work_book.active
            self.work_sheet.title = "备案信息" 
            self.work_sheet.freeze_panes = 'A2' # 冻结窗格
            col_title_list = ['域名主办方', '域名', '备案许可证号', '网站备案号', '域名类型', '网站前置审批项', '是否限制接入', '审核通过日期']
            col_width = {'A': 45, 'B': 30, 'C': 22, 'D': 24, 'E': 9, 'F': 41, 'G': 13, 'H': 21}
            for i in range(8):
                self.work_sheet.cell(1, i + 1).value = col_title_list[i] # 写入表头
            for key, value in col_width.items():
                self.work_sheet.column_dimensions[key].width = value # 设定列宽
            self.start = 0
            self.after_title = 2
            self.total_row = list_row
        self.write_data()
        
    def write_data(self):
        print(f"{' 正在录入信息到表格 ':-^50}\n")
        for i in range(self.start, self.total_row):
            for j in range(8):
                self.work_sheet.cell(i + self.after_title, j + 1).value = self.domain_result_list[i - self.start][j] # 写入数据，对应domain_reuslt_list的数据，结构是：[(1,2,3,4,5,6,7,8),(),()……]
        for row in range(self.work_sheet.max_row):
            for col in range(self.work_sheet.max_column):
                self.work_sheet.cell(row + 1, col + 1).alignment = Alignment(horizontal='center', vertical='center') # 居中对齐
        try:
            self.work_book.save(self.excel_file_path)
        except PermissionError:
            return print("** 备案信息登记表格已打开，无法写入文件。如需写入，请关闭文件后重新执行！ **\n")
        return print(f"{' 查询结果保存在：' + str(self.excel_file_path) + ' ':*^47}\n")

if __name__ == '__main__':
   """
   基本流程：获取Cookie、获取Token、获取大小图片、点选、得到加密坐标回传得到Sign、获得查询结果，各类可预期的报错，返回为-1
   说明：在我的测试环境里，不用验证图片也可以获取到信息（此前必须过了验证才行）。可以自行将270到272行代码注释掉进行测试
   """
   os.environ['no_proxy'] = '*' # 不通过代理，避免SSL报错
   query_target = query_info() # 输入过滤
   requirement_info = get_require_info() # 获得Cookie、Token
   pic_check = image_tool() # 图片处理以及手动图片点选
   save_data = excel_tool() # 
   print(f"\n{' 项目地址：https://github.com/wongzeon/ICP-Checker ':-^54}\n")
   print(f"{' V2.2.0 2023/11/23 仅用于学习研究，不得用于商业/非法用途 ':-^41}")
   requirement_info.get_cookie() 
   token = requirement_info.get_token() 
   exe_num = 0
   while exe_num <= 2:
        if query_target.regular_input() != -1: 
            if pic_check.check_image() == -1:
                print(f"\n{' 验证失败，请重新验证 ':*^49}")      
                pic_check.check_image()
            if query_target.query_information() != -1: 
                save_data.set_format()
        exe_num += 1 # 控制循环次数
